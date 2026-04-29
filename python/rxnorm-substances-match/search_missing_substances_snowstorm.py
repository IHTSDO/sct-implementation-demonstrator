#!/usr/bin/env python3
"""Count or search missing RxNorm substances in Snowstorm.

The script reads the second sheet in the workbook by default and uses the
``RxNorm Ingredient`` column as the search term source. It defaults to
count-only mode so the input set can be checked before issuing requests.
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook


DEFAULT_WORKBOOK = Path(
    "~/Downloads/missing-substances-rxnorm-snomed-2503 (1).xlsx"
)
DEFAULT_OUTPUT = Path("tmp/substances/snowstorm-substance-search-results.jsonl")
DEFAULT_XLSX_OUTPUT = Path("tmp/substances/snowstorm-substance-first-matches.xlsx")
BASE_URL = (
    "https://browser.ihtsdotools.org/snowstorm/snomed-ct/multisearch/descriptions"
)


def normalize_term(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_terms(workbook_path: Path, sheet_index: int, term_column_name: str) -> tuple[str, list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[sheet_index]
    except IndexError as exc:
        raise SystemExit(
            f"Workbook has {len(workbook.worksheets)} sheets; sheet index {sheet_index + 1} is unavailable."
        ) from exc

    rows = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise SystemExit(f"Sheet {worksheet.title!r} is empty.") from exc

    normalized_header = [normalize_term(value).lower() for value in header]
    try:
        term_column_index = normalized_header.index(term_column_name.lower())
    except ValueError as exc:
        available = ", ".join(normalize_term(value) for value in header if normalize_term(value))
        raise SystemExit(
            f"Column {term_column_name!r} was not found in sheet {worksheet.title!r}. "
            f"Available columns: {available}"
        ) from exc

    terms = []
    for row in rows:
        if term_column_index >= len(row):
            continue
        term = normalize_term(row[term_column_index])
        if term:
            terms.append(term)

    return worksheet.title, terms


def build_search_url(term: str, limit: int) -> str:
    query = {
        "offset": 0,
        "limit": limit,
        "active": "true",
        "conceptActive": "true",
        "term": term,
    }
    return f"{BASE_URL}?{urlencode(query)}"


def search_term(term: str, limit: int, timeout_seconds: int) -> dict[str, Any]:
    url = build_search_url(term, limit)
    request = Request(url, headers={"Accept": "application/json", "User-Agent": "sct-search-script/1.0"})
    with urlopen(request, timeout=timeout_seconds) as response:
        body = response.read().decode("utf-8")
    payload = json.loads(body)
    return {
        "term": term,
        "url": url,
        "status": "ok",
        "result": payload,
    }


def first_match_row(source_term: str, result: dict[str, Any]) -> list[Any]:
    if result["status"] != "ok":
        return [
            source_term,
            "error",
            None,
            None,
            None,
            None,
            None,
            None,
            result.get("error"),
        ]

    payload = result["result"]
    items = payload.get("items", [])
    total = payload.get("total", len(items))
    if not items:
        return [source_term, total, None, None, None, None, None, None, None]

    first_item = items[0]
    concept = first_item.get("concept", {})
    fsn = concept.get("fsn") or {}
    preferred_term = concept.get("pt") or {}
    return [
        source_term,
        total,
        concept.get("conceptId") or concept.get("id"),
        first_item.get("term"),
        preferred_term.get("term"),
        fsn.get("term"),
        first_item.get("branchPath"),
        first_item.get("languageCode"),
        None,
    ]


def count_duplicates(terms: list[str]) -> int:
    seen = set()
    duplicates = 0
    for term in terms:
        key = term.casefold()
        if key in seen:
            duplicates += 1
        else:
            seen.add(key)
    return duplicates


def print_count_summary(workbook_path: Path, sheet_name: str, terms: list[str]) -> None:
    duplicate_count = count_duplicates(terms)
    unique_count = len({term.casefold() for term in terms})

    print(f"Workbook: {workbook_path}")
    print(f"Sheet: {sheet_name}")
    print("Term column: RxNorm Ingredient")
    print(f"Non-empty terms: {len(terms)}")
    print(f"Unique terms, case-insensitive: {unique_count}")
    print(f"Duplicate rows, case-insensitive: {duplicate_count}")
    print("First 10 terms:")
    for index, term in enumerate(terms[:10], start=1):
        print(f"  {index}. {term}")


def run_searches(
    terms: list[str],
    output_path: Path,
    xlsx_output_path: Path,
    delay_seconds: float,
    limit: int,
    timeout_seconds: int,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "First matches"
    worksheet.append(
        [
            "Source Term",
            "Result Count",
            "Matched Code",
            "Matched Description Term",
            "Matched Preferred Term",
            "Matched FSN",
            "Branch Path",
            "Language Code",
            "Error",
        ]
    )

    with output_path.open("w", encoding="utf-8") as output_file:
        for index, term in enumerate(terms, start=1):
            print(f"[{index}/{len(terms)}] Searching: {term}", flush=True)
            try:
                result = search_term(term, limit, timeout_seconds)
            except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
                result = {
                    "term": term,
                    "url": build_search_url(term, limit),
                    "status": "error",
                    "error": str(exc),
                }

            output_file.write(json.dumps(result, ensure_ascii=False) + "\n")
            output_file.flush()
            worksheet.append(first_match_row(term, result))

            if index < len(terms):
                time.sleep(delay_seconds)

    workbook.save(xlsx_output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Count or sequentially search RxNorm Ingredient terms in Snowstorm."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet-index", type=int, default=2, help="1-based worksheet index. Default: 2.")
    parser.add_argument("--term-column", default="RxNorm Ingredient")
    parser.add_argument("--run", action="store_true", help="Run the Snowstorm searches after counting.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--xlsx-output", type=Path, default=DEFAULT_XLSX_OUTPUT)
    parser.add_argument("--delay-seconds", type=float, default=1.0)
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--timeout-seconds", type=int, default=30)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    sheet_name, terms = get_terms(args.workbook, args.sheet_index - 1, args.term_column)
    print_count_summary(args.workbook, sheet_name, terms)

    if not args.run:
        print("Searches were not run. Add --run to query Snowstorm sequentially.")
        return 0

    run_searches(terms, args.output, args.xlsx_output, args.delay_seconds, args.limit, args.timeout_seconds)
    print(f"Search results written to: {args.output}")
    print(f"First matches written to: {args.xlsx_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
